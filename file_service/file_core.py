from openpyxl import load_workbook
from utils import DatabaseService
from file_service import get_path

db = DatabaseService()


async def read_file(file_path: str, subject_id: int):
    try:
        sheet = load_workbook(await get_path(file_name=file_path)).active
        if await db.get_question(subject_id=subject_id, text=next(sheet.iter_rows(values_only=True))[1]):
            print("Malumot bazasida savol mavjud")
            return None
        else:
            print("Malumot bazasida savol mavjud emas")
            for row in sheet.iter_rows(values_only=True):
                await db.add_question(subject_id=subject_id, question=row[1], option1=str(row[2]), option2=str(row[3]), option3=str(row[4]),
                                      option4=str(row[5]))
    except Exception as e:
        print(e)
        return None


async def read_user_info(passport: str):
    sheet = load_workbook(await get_path(file_name='user_info.xlsx')).active
    for row in sheet.iter_rows(values_only=True):
        if str(row[1]) == str(passport):
            return row
